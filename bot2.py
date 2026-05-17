from aiogram import Bot, Dispatcher, types
from aiogram.filters import Command
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
import asyncio
import logging
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

API_TOKEN = ''
bot = Bot(token=API_TOKEN)
dp = Dispatcher()

user_data = {}

spisok = InlineKeyboardMarkup(inline_keyboard=[
    [InlineKeyboardButton(text="Спринт", callback_data="sprint")],
    [InlineKeyboardButton(text="Средние дистанции", callback_data="middle")],
    [InlineKeyboardButton(text="Длинные дистанции", callback_data="long")],
    [InlineKeyboardButton(text="Эстафета", callback_data="relay")]
])

spisok2 = InlineKeyboardMarkup(inline_keyboard=[
    [InlineKeyboardButton(text="Да", callback_data="yes")],
    [InlineKeyboardButton(text="Нет", callback_data="no")]
])

@dp.message(Command("start"))
async def start(message: types.Message):
    await message.answer("Привет!\n\nЯ - Чат-бот записи на марафон.\n\nЕсли ты хочешь записаться на марафон жми /go!")

@dp.message(Command("go"))
async def reg(message: types.Message):
    await message.answer("Отлично!\n\nДля записи на марафон требуется ваше твоё ФИО и тип забега.\n\nВведите своё ФИО одним сообщением:\nПример: Стольников Артём Вячеславович")

@dp.message()
async def get_name(message: types.Message):
    user_id = message.from_user.id
    proverka = message.text.split()
    if user_id not in user_data:
        if len(proverka) != 3:
            await message.answer("Ошибка!\n\nВы ввели ненастоящее ФИО.\n\nВведите ваше настоящее ФИО.")
            return
        user_data[user_id] = {"name": message.text}
        await message.answer(f"Приятно познакомиться, {message.text}!\n\nТеперь выберите тип забега: *тык*", reply_markup=spisok)

@dp.callback_query(lambda c: c.data in ["sprint", "middle", "long", "relay"])
async def race_choice(callback: types.CallbackQuery):
    user_id = callback.from_user.id
    race_map = {
        "sprint": "Спринт",
        "middle": "Средние дистанции",
        "long": "Длинные дистанции",
        "relay": "Эстафета"
    }
    user_data[user_id]["race"] = race_map[callback.data]
    await callback.message.edit_reply_markup(reply_markup=None)
    fio = user_data[user_id]["name"]
    race = user_data[user_id]["race"]
    await callback.message.answer(
        f"Проверьте ваши данные:\n\nВы - {fio},\nВыбранный тип марафона - {race}?",
        reply_markup=spisok2
    )
    await callback.answer()

@dp.callback_query(lambda c: c.data in ["yes", "no"])
async def confirm_choice(callback: types.CallbackQuery):
    user_id = callback.from_user.id
    fio = user_data.get(user_id, {}).get("name", "Неизвестно")
    race = user_data.get(user_id, {}).get("race", "Неизвестно")
    save(fio, race)
    if callback.data == "yes":
        await callback.message.edit_text(f"Поздравляю! Вы записаны на марафон.\n\nВаши данные:\nФИО: {fio}\nТип марафона: {race}")
    else:
        await callback.message.edit_text("Запись отменена.\nНапишите /go, чтобы попробовать снова.")
    await callback.answer()
    user_data.pop(user_id, None)

def save(fio:str, race:str, filename:str="marathon.xlsx"):
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Регистрации"
        ws.append(["ФИО", "Тип забега", "Дата регистрации"])
    current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([fio, race, current_date])
    wb.save(filename)

async def main():
    logging.basicConfig(level=logging.INFO)
    print("Бот запущен...")
    await bot.delete_webhook(drop_pending_updates=True)
    await dp.start_polling(bot)

if __name__ == '__main__':
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("Бот выключен")
